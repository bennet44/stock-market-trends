# -*- coding: utf-8 -*-
"""將 FCN 資料夾兩張手寫筆記 (28236.jpg / 28237.jpg) 整理成 Excel 表格。

資料來源為 src/fcn_records.py（18 筆 FCN 實單，單一真相來源）。期初價一律改用
「定價日的實際收盤價」（透過 src/price_fetch 即時抓取），並保留手寫原記供稽核。
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

from src import fcn_records as fr


def _pct(frac, label=""):
    return f"{frac * 100:.4g}%{label}" if frac is not None else "—"


def _ki_type(r):
    style = "每日" if r.ki_style == "continuous" else "到期"
    return f"{r.ki_label}（{style}）" if r.ki_label != "—" else f"—（{style}）"


def _lines(tickers, prices):
    out = []
    for t, p in zip(tickers, prices):
        out.append(f"{t} {p:.2f}" if isinstance(p, (int, float)) else f"{t} —")
    return "\n".join(out)


print("抓取各標的定價日實際收盤價…")
records = fr.attach_actual_prices(verbose=True)

headers = ["序", "交易日", "定價日", "標的 / 期初價（實際收盤）", "手寫原記",
           "天期(月)", "配息率(年化)", "KO%", "執行價%", "下限(KI)%", "KI類型", "張數", "結構 / 備註"]

wb = Workbook()
ws = wb.active
ws.title = "FCN 記錄"

title_font = Font(name="微軟正黑體", size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
cell_font = Font(name="微軟正黑體", size=10)
audit_font = Font(name="微軟正黑體", size=9, color="808080")
note_font = Font(name="微軟正黑體", size=9, italic=True, color="808080")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

ncol = len(headers)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
c = ws.cell(row=1, column=1,
            value="FCN（固定配息連動債）記錄表　—　期初價＝2026 定價日實際收盤價（手寫原記並列供稽核）")
c.font = title_font; c.fill = title_fill; c.alignment = center
ws.row_dimensions[1].height = 26

hr = 2
for j, h in enumerate(headers, 1):
    c = ws.cell(row=hr, column=j, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
ws.row_dimensions[hr].height = 30

alt = PatternFill("solid", fgColor="EAF1FB")
for i, r in enumerate(records):
    rr = hr + 1 + i
    date_disp = r.trade_date[5:] + ("（不清）" if r.date_uncertain else "")
    pricing_disp = r.pricing_date[5:] + ("（推估）" if r.date_uncertain else "")
    values = [
        r.id, date_disp, pricing_disp,
        _lines(r.tickers, r.actual),
        _lines(r.tickers, r.handwritten),
        r.tenor_months, _pct(r.coupon), _pct(r.ko), _pct(r.strike), _pct(r.ki),
        _ki_type(r), f"{r.zhang}張", r.notes,
    ]
    for j, val in enumerate(values, 1):
        c = ws.cell(row=rr, column=j, value=val)
        c.font = audit_font if j == 5 else cell_font
        c.border = border
        c.alignment = left if j in (4, 5, 13) else center
        if i % 2 == 1:
            c.fill = alt
    ws.row_dimensions[rr].height = 14 * len(r.tickers) + 6

nr = hr + 1 + len(records) + 1
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=ncol)
note = ("說明：期初價採「定價日實際收盤價」；「手寫原記」保留筆記原數字供對照，兩者差異多為交易日與"
        "定價日相隔數日的行情漂移。少數手寫誤差較大（如第11筆 AMD 手寫 159.85→實際約 532、第9筆 GOOGL"
        " 手寫 262→實際約 364）為字跡誤讀，分析一律以實際收盤為準。KO=提前出場、執行價=轉換價、下限=KI"
        " 障礙，皆為期初價百分比。KI 類型：AKI＝每日觀察、飛KI/空白＝到期觀察。第9筆日期不清，依期初價回推。"
        "第17、18筆手寫未定價，期初價取交易日實際收盤。")
c = ws.cell(row=nr, column=1, value=note)
c.font = note_font; c.alignment = left
ws.row_dimensions[nr].height = 78

widths = [4, 12, 12, 22, 16, 8, 11, 8, 9, 9, 14, 7, 26]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A3"

# ---- 統計摘要 ----
ws2 = wb.create_sheet("統計摘要")


def put(rr, cc, v, font=None, fill=None, align=None):
    cell = ws2.cell(row=rr, column=cc, value=v)
    cell.font = font or cell_font
    if fill:
        cell.fill = fill
    cell.alignment = align or center
    cell.border = border
    return cell


ws2.merge_cells("A1:B1")
t = ws2.cell(row=1, column=1, value="FCN 統計摘要")
t.font = title_font; t.fill = title_fill; t.alignment = center
ws2.row_dimensions[1].height = 24

uc = Counter()
for r in records:
    for t_ in r.tickers:
        uc[t_] += 1

row = 3
put(row, 1, "標的", hdr_font, hdr_fill); put(row, 2, "出現次數", hdr_font, hdr_fill)
for tk, n in uc.most_common():
    row += 1
    put(row, 1, tk, cell_font, None, left); put(row, 2, n)

kc = Counter(r.ki_label if r.ki_label != "—" else "未註記" for r in records)
tc = Counter(f"{r.tenor_months}月" for r in records)

start = row + 2
put(start, 1, "KI 類型", hdr_font, hdr_fill); put(start, 2, "筆數", hdr_font, hdr_fill)
for k, n in kc.most_common():
    start += 1
    put(start, 1, k, cell_font, None, left); put(start, 2, n)

start += 2
put(start, 1, "天期", hdr_font, hdr_fill); put(start, 2, "筆數", hdr_font, hdr_fill)
for k, n in sorted(tc.items(), key=lambda x: int(x[0][:-1])):
    start += 1
    put(start, 1, k, cell_font, None, left); put(start, 2, n)

start += 2
put(start, 1, "總筆數", hdr_font, hdr_fill); put(start, 2, len(records))
start += 1
put(start, 1, "張數合計", cell_font, None, left); put(start, 2, f"{sum(r.zhang for r in records)} 張")
start += 1
put(start, 1, "平均配息率(年化)", cell_font, None, left)
put(start, 2, f"{sum(r.coupon for r in records) / len(records) * 100:.2f}%")

for j, w in enumerate([18, 10], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w

out = Path(__file__).resolve().parent / "FCN_記錄整理.xlsx"
wb.save(out)
print("saved:", out)
